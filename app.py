import os
import sqlite3
import smtplib
from datetime import datetime, date, timedelta
from email.message import EmailMessage
from io import BytesIO

import psycopg2
import psycopg2.extras
from apscheduler.schedulers.background import BackgroundScheduler
from flask import Flask, request, redirect, render_template, jsonify, send_file, abort
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)

ACCESS_CODE = os.environ.get("ACCESS_CODE", "h2otech")
DATABASE_URL = os.environ.get("DATABASE_URL")

EMAIL_TO = os.environ.get("EMAIL_TO", "")
EMAIL_FROM = os.environ.get("EMAIL_FROM", "")
SMTP_HOST = os.environ.get("SMTP_HOST", "")
SMTP_PORT = int(os.environ.get("SMTP_PORT", "587"))
SMTP_USER = os.environ.get("SMTP_USER", "")
SMTP_PASSWORD = os.environ.get("SMTP_PASSWORD", "")

LOCAL_DB = "colis.db"


def is_postgres():
    return bool(DATABASE_URL and DATABASE_URL.startswith(("postgres://", "postgresql://")))


def get_conn():
    if is_postgres():
        return psycopg2.connect(DATABASE_URL)
    return sqlite3.connect(LOCAL_DB)


def init_db():
    conn = get_conn()
    cur = conn.cursor()
    if is_postgres():
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS colis (
                id SERIAL PRIMARY KEY,
                client TEXT NOT NULL,
                adresse TEXT NOT NULL,
                contenu TEXT NOT NULL,
                statut TEXT NOT NULL DEFAULT 'A FAIRE',
                created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                done_at TIMESTAMP NULL
            )
            """
        )
    else:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS colis (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                client TEXT NOT NULL,
                adresse TEXT NOT NULL,
                contenu TEXT NOT NULL,
                statut TEXT NOT NULL DEFAULT 'A FAIRE',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP,
                done_at TEXT NULL
            )
            """
        )
    conn.commit()
    conn.close()


def require_code():
    code = request.values.get("code", "")
    if code != ACCESS_CODE:
        abort(403)


def rows_to_dicts(cur):
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, row)) for row in cur.fetchall()]


def get_colis(include_done=True, month_start=None, month_end=None):
    conn = get_conn()
    cur = conn.cursor()
    params = []
    where = []
    if not include_done:
        where.append("statut <> 'FAIT'")
    if month_start and month_end:
        where.append("created_at >= %s" if is_postgres() else "created_at >= ?")
        params.append(month_start)
        where.append("created_at < %s" if is_postgres() else "created_at < ?")
        params.append(month_end)
    sql = "SELECT id, client, adresse, contenu, statut, created_at, done_at FROM colis"
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY CASE WHEN statut='A FAIRE' THEN 0 ELSE 1 END, created_at DESC"
    cur.execute(sql, params)
    data = rows_to_dicts(cur)
    conn.close()
    return data


def month_bounds(target=None):
    target = target or date.today()
    start = date(target.year, target.month, 1)
    if target.month == 12:
        end = date(target.year + 1, 1, 1)
    else:
        end = date(target.year, target.month + 1, 1)
    return start, end


@app.route("/", methods=["GET", "POST"])
def index():
    code = request.values.get("code", "")
    if code != ACCESS_CODE:
        return render_template("login.html")

    if request.method == "POST":
        client = request.form.get("client", "").strip()
        adresse = request.form.get("adresse", "").strip()
        contenu = request.form.get("contenu", "").strip()
        if client and adresse and contenu:
            conn = get_conn()
            cur = conn.cursor()
            ph = "%s" if is_postgres() else "?"
            cur.execute(
                f"INSERT INTO colis (client, adresse, contenu, statut) VALUES ({ph}, {ph}, {ph}, 'A FAIRE')",
                (client, adresse, contenu),
            )
            conn.commit()
            conn.close()
        return redirect(f"/?code={ACCESS_CODE}")

    colis = get_colis(include_done=True)
    return render_template("index.html", colis=colis, code=ACCESS_CODE)


@app.route("/api/colis")
def api_colis():
    return jsonify(get_colis(include_done=True))


@app.route("/ecran")
def ecran():
    return render_template("ecran.html")


@app.route("/statut/<int:colis_id>/<statut>", methods=["POST", "GET"])
def statut(colis_id, statut):
    require_code()
    statut = statut.upper()
    if statut not in ["A FAIRE", "FAIT"]:
        abort(400)
    conn = get_conn()
    cur = conn.cursor()
    done_at = datetime.now().isoformat(timespec="seconds") if statut == "FAIT" else None
    ph = "%s" if is_postgres() else "?"
    cur.execute(f"UPDATE colis SET statut={ph}, done_at={ph} WHERE id={ph}", (statut, done_at, colis_id))
    conn.commit()
    conn.close()
    return redirect(f"/?code={ACCESS_CODE}")


@app.route("/delete/<int:colis_id>", methods=["POST", "GET"])
def delete(colis_id):
    require_code()
    conn = get_conn()
    cur = conn.cursor()
    ph = "%s" if is_postgres() else "?"
    cur.execute(f"DELETE FROM colis WHERE id={ph}", (colis_id,))
    conn.commit()
    conn.close()
    return redirect(f"/?code={ACCESS_CODE}")


@app.route("/historique")
def historique():
    require_code()
    colis = get_colis(include_done=True)
    return render_template("historique.html", colis=colis, code=ACCESS_CODE)


def build_excel(rows, title="Suivi colis"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Envois colis"
    ws.append([title])
    ws.merge_cells("A1:F1")
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.append(["Date création", "Client", "Adresse", "Colis / contenu", "Statut", "Date fait"])
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
    for r in rows:
        ws.append([r.get("created_at"), r.get("client"), r.get("adresse"), r.get("contenu"), r.get("statut"), r.get("done_at")])
    widths = [22, 25, 40, 55, 15, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    for row in ws.iter_rows(min_row=3):
        status = row[4].value
        fill = PatternFill("solid", fgColor="C6EFCE" if status == "FAIT" else "FFC7CE")
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.fill = fill
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


@app.route("/export/monthly")
def export_monthly():
    require_code()
    start, end = month_bounds()
    rows = get_colis(include_done=True, month_start=start.isoformat(), month_end=end.isoformat())
    bio = build_excel(rows, f"Envois colis - {start.strftime('%m/%Y')}")
    return send_file(bio, as_attachment=True, download_name=f"envois_colis_{start.strftime('%Y_%m')}.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


def send_monthly_report(target=None):
    if not all([EMAIL_TO, EMAIL_FROM, SMTP_HOST, SMTP_USER, SMTP_PASSWORD]):
        return False, "Configuration email incomplète"
    start, end = month_bounds(target)
    rows = get_colis(include_done=True, month_start=start.isoformat(), month_end=end.isoformat())
    bio = build_excel(rows, f"Envois colis - {start.strftime('%m/%Y')}")

    msg = EmailMessage()
    msg["Subject"] = f"Liste des envois colis - {start.strftime('%m/%Y')}"
    msg["From"] = EMAIL_FROM
    msg["To"] = EMAIL_TO
    msg.set_content(f"Bonjour,\n\nCi-joint la liste des envois colis du mois {start.strftime('%m/%Y')}.\n\nNombre d'envois : {len(rows)}\n")
    msg.add_attachment(bio.getvalue(), maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=f"envois_colis_{start.strftime('%Y_%m')}.xlsx")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(SMTP_USER, SMTP_PASSWORD)
        smtp.send_message(msg)
    return True, "Rapport envoyé"


@app.route("/send_report")
def send_report():
    require_code()
    ok, message = send_monthly_report()
    return message, (200 if ok else 500)


def schedule_jobs():
    scheduler = BackgroundScheduler(timezone="Europe/Paris")
    # Tous les jours à 18h, on vérifie si on est le dernier jour du mois.
    def maybe_send():
        today = date.today()
        tomorrow = today + timedelta(days=1)
        if tomorrow.month != today.month:
            send_monthly_report(today)
    scheduler.add_job(maybe_send, "cron", hour=18, minute=0)
    scheduler.start()


init_db()
schedule_jobs()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=False)
